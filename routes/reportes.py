import io
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, session
from database import get_db
from utils import login_required, admin_required
from datetime import datetime

reportes_bp = Blueprint('reportes', __name__)


@reportes_bp.route('/reportes')
@admin_required
def index():
    db = get_db()
    comercio_id = session['comercio_id']
    comercio = db.execute('SELECT * FROM comercios WHERE id = ?', (comercio_id,)).fetchone()
    
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')
    
    # Valores de búsqueda por defecto (los últimos 30 días)
    if not fecha_desde:
        from datetime import timedelta
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')

    # 1. Ventas
    ventas_total = db.execute('''
        SELECT COALESCE(SUM(total), 0) FROM ventas 
        WHERE comercio_id = ? AND anulada = 0 AND DATE(fecha) BETWEEN ? AND ?
    ''', (comercio_id, fecha_desde, fecha_hasta)).fetchone()[0]

    # 2. Compras
    compras_total = db.execute('''
        SELECT COALESCE(SUM(total), 0) FROM compras 
        WHERE comercio_id = ? AND DATE(fecha) BETWEEN ? AND ?
    ''', (comercio_id, fecha_desde, fecha_hasta)).fetchone()[0]

    # 3. Gastos
    gastos_total = db.execute('''
        SELECT COALESCE(SUM(monto), 0) FROM gastos 
        WHERE comercio_id = ? AND DATE(fecha) BETWEEN ? AND ?
    ''', (comercio_id, fecha_desde, fecha_hasta)).fetchone()[0]

    ganancia_neta = ventas_total - compras_total - gastos_total
    
    estado_resultados = {
        'ventas': ventas_total,
        'compras': compras_total,
        'gastos': gastos_total,
        'ganancia_neta': ganancia_neta,
        'desde': fecha_desde,
        'hasta': fecha_hasta
    }

    # Métricas de valoración de inventario si aplica (Ferretería / General)
    inventario_val = None
    if comercio['tipo'] in ('ferreteria', 'general'):
        stats = db.execute('''
            SELECT 
                SUM(stock * costo) as total_costo,
                SUM(stock * precio) as total_venta,
                SUM(stock * (precio - costo)) as total_ganancia
            FROM productos
            WHERE activo = 1 AND stock IS NOT NULL AND costo IS NOT NULL AND precio IS NOT NULL AND comercio_id = ?
        ''', (comercio_id,)).fetchone()
        inventario_val = dict(stats) if stats['total_costo'] is not None else None
        
    return render_template('reportes.html', api_key=comercio['api_key'], inventario_val=inventario_val, estado_resultados=estado_resultados)


@reportes_bp.route('/reportes/excel')
@admin_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Side, Border

    db = get_db()
    comercio_id = session['comercio_id']
    comercio = db.execute('SELECT * FROM comercios WHERE id = ?', (comercio_id,)).fetchone()
    
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    query = '''
        SELECT v.id, v.fecha, v.total, v.tipo_pago,
               COALESCE(c.nombre, '-') as cliente,
               u.nombre as vendedor,
               GROUP_CONCAT(p.nombre || ' x' || dv.cantidad, ', ') as productos
        FROM ventas v
        JOIN usuarios u ON v.usuario_id = u.id
        LEFT JOIN clientes c ON v.cliente_id = c.id
        LEFT JOIN detalle_ventas dv ON dv.venta_id = v.id
        LEFT JOIN productos p ON dv.producto_id = p.id
        WHERE v.anulada = 0 AND v.comercio_id = ?
    '''
    params = [comercio_id]
    if fecha_desde:
        query += ' AND DATE(v.fecha) >= ?'; params.append(fecha_desde)
    if fecha_hasta:
        query += ' AND DATE(v.fecha) <= ?'; params.append(fecha_hasta)
    query += ' GROUP BY v.id ORDER BY v.fecha DESC'

    ventas = db.execute(query, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Estilos
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Fecha', 'Hora', 'Productos', 'Total ($)', 'Tipo Pago', 'Cliente', 'Vendedor']
    col_widths = [6, 14, 10, 45, 14, 16, 20, 16]

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{comercio['logo_emoji']} {comercio['nombre']} — Reporte de Ventas"
    ws['A1'].font = Font(bold=True, size=14, color='7C3AED')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Subtítulo
    ws.merge_cells('A2:H2')
    periodo = f'{fecha_desde or "Inicio"} al {fecha_hasta or "Hoy"}'
    ws['A2'] = f'Período: {periodo} | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])  # fila vacía

    # Encabezados
    ws.append(headers)
    header_row = ws.max_row
    for col, (cell, width) in enumerate(zip(ws[header_row], col_widths), 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    # Datos
    total_general = 0
    for v in ventas:
        fecha_dt = v['fecha'][:10] if v['fecha'] else ''
        hora = v['fecha'][11:16] if v['fecha'] and len(v['fecha']) > 10 else ''
        ws.append([
            v['id'], fecha_dt, hora,
            v['productos'] or '',
            v['total'], v['tipo_pago'],
            v['cliente'], v['vendedor']
        ])
        total_general += v['total']
        row = ws.max_row
        for cell in ws[row]:
            cell.border = border
            if cell.column == 5:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    # Total
    ws.append([])
    total_row = ws.max_row + 1
    ws.append(['', '', '', 'TOTAL GENERAL', total_general, '', '', ''])
    tr = ws.max_row
    ws[f'D{tr}'].font = Font(bold=True)
    ws[f'E{tr}'].font = Font(bold=True, color='7C3AED')
    ws[f'E{tr}'].number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'reporte_ventas_{comercio["tipo"]}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      as_attachment=True, download_name=nombre_archivo)


@reportes_bp.route('/reportes/pdf')
@admin_required
def exportar_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    db = get_db()
    comercio_id = session['comercio_id']
    comercio = db.execute('SELECT * FROM comercios WHERE id = ?', (comercio_id,)).fetchone()
    
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    query = '''
        SELECT v.id, v.fecha, v.total, v.tipo_pago,
               COALESCE(c.nombre, '-') as cliente, u.nombre as vendedor
        FROM ventas v
        JOIN usuarios u ON v.usuario_id = u.id
        LEFT JOIN clientes c ON v.cliente_id = c.id
        WHERE v.anulada = 0 AND v.comercio_id = ?
    '''
    params = [comercio_id]
    if fecha_desde:
        query += ' AND DATE(v.fecha) >= ?'; params.append(fecha_desde)
    if fecha_hasta:
        query += ' AND DATE(v.fecha) <= ?'; params.append(fecha_hasta)
    query += ' ORDER BY v.fecha DESC'
    
    ventas = db.execute(query, params).fetchall()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle('title', fontSize=18, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#7C3AED'), spaceAfter=6)
    elements.append(Paragraph(f"{comercio['logo_emoji']} {comercio['nombre']} — Reporte de Ventas", title_style))

    sub_style = ParagraphStyle('sub', fontSize=10, textColor=colors.grey, spaceAfter=14)
    periodo = f'{fecha_desde or "Inicio"} al {fecha_hasta or "Hoy"}'
    elements.append(Paragraph(f'Período: {periodo}  |  Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_style))

    # Tabla
    data = [['#', 'Fecha', 'Hora', 'Total', 'Tipo Pago', 'Cliente', 'Vendedor']]
    total_general = 0
    for v in ventas:
        fecha_str = v['fecha'][:10] if v['fecha'] else ''
        hora_str = v['fecha'][11:16] if v['fecha'] and len(v['fecha']) > 10 else ''
        total_general += v['total']
        data.append([
            str(v['id']), fecha_str, hora_str,
            f"${v['total']:,.0f}".replace(',', '.'),
            v['tipo_pago'], v['cliente'], v['vendedor']
        ])

    data.append(['', '', '', f"TOTAL: ${total_general:,.0f}".replace(',', '.'), '', '', ''])

    purple = colors.HexColor('#7C3AED')
    col_widths_pdf = [1.5*cm, 2.5*cm, 1.8*cm, 3*cm, 3*cm, 4*cm, 3.5*cm]
    table = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), purple),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor('#F5F0FF'), colors.white]),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (3,-1), (3,-1), purple),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    nombre_archivo = f'reporte_ventas_{comercio["tipo"]}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name=nombre_archivo)
